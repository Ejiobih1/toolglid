"""
Advanced PDF to Excel Converter
High-quality table extraction from PDFs using Camelot.

Uses:
- Camelot for accurate table detection and extraction
- openpyxl for Excel file creation
- Supports both lattice (bordered) and stream (borderless) tables

Author: ToolGlid
"""

import camelot
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import io
import tempfile
import os


class PDFToExcelConverter:
    """
    High-quality PDF to Excel converter using Camelot.

    Features:
    - Automatic table detection (lattice and stream modes)
    - Multi-page PDF support
    - Each table can go to separate sheet or all in one
    - Preserves cell structure and basic formatting
    - Auto-adjusts column widths
    """

    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)

        # Style definitions
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _extract_tables_lattice(self, pages: str = 'all') -> List:
        """Extract tables using lattice method (for bordered tables)"""
        try:
            tables = camelot.read_pdf(
                self.pdf_path,
                pages=pages,
                flavor='lattice',
                strip_text='\n'
            )
            return tables
        except Exception as e:
            print(f"Lattice extraction failed: {e}")
            return []

    def _extract_tables_stream(self, pages: str = 'all') -> List:
        """Extract tables using stream method (for borderless tables)"""
        try:
            tables = camelot.read_pdf(
                self.pdf_path,
                pages=pages,
                flavor='stream',
                strip_text='\n',
                edge_tol=50,
                row_tol=10
            )
            return tables
        except Exception as e:
            print(f"Stream extraction failed: {e}")
            return []

    def _auto_detect_tables(self, pages: str = 'all') -> List:
        """
        Auto-detect tables trying lattice first, then stream.
        Returns the method that finds more/better tables.
        """
        # Try lattice first (works best for bordered tables)
        lattice_tables = self._extract_tables_lattice(pages)

        # Try stream (works for borderless tables)
        stream_tables = self._extract_tables_stream(pages)

        # Calculate quality scores
        lattice_score = sum(t.accuracy for t in lattice_tables) if lattice_tables else 0
        stream_score = sum(t.accuracy for t in stream_tables) if stream_tables else 0

        # Return the better result
        if lattice_score >= stream_score and len(lattice_tables) > 0:
            print(f"Using lattice method: {len(lattice_tables)} tables, score: {lattice_score:.1f}")
            return list(lattice_tables)
        elif len(stream_tables) > 0:
            print(f"Using stream method: {len(stream_tables)} tables, score: {stream_score:.1f}")
            return list(stream_tables)
        else:
            # Return whichever has tables
            return list(lattice_tables) if lattice_tables else list(stream_tables)

    def _add_table_to_sheet(self, table, sheet, start_row: int = 1) -> int:
        """
        Add a Camelot table to an Excel sheet.
        Returns the next available row.
        """
        df = table.df

        # Write data
        for row_idx, row in enumerate(df.values):
            excel_row = start_row + row_idx
            for col_idx, value in enumerate(row):
                cell = sheet.cell(row=excel_row, column=col_idx + 1)
                cell.value = str(value).strip() if value else ""
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Style first row as header
                if row_idx == 0:
                    cell.font = self.header_font
                    cell.fill = self.header_fill

        # Auto-adjust column widths
        for col_idx in range(len(df.columns)):
            column_letter = get_column_letter(col_idx + 1)
            max_length = 0
            for row in df.values:
                try:
                    cell_value = str(row[col_idx]) if col_idx < len(row) else ""
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            sheet.column_dimensions[column_letter].width = max(adjusted_width, 10)

        return start_row + len(df) + 2  # Return next row with spacing

    def convert(self, output_path: str = None, separate_sheets: bool = True,
                pages: str = 'all', flavor: str = 'auto') -> str:
        """
        Convert PDF tables to Excel.

        Args:
            output_path: Output .xlsx path. If None, uses input name with .xlsx extension.
            separate_sheets: If True, each table goes to a separate sheet.
                           If False, all tables go to one sheet.
            pages: Page numbers to extract ('all', '1', '1,2,3', '1-5')
            flavor: 'lattice' for bordered tables, 'stream' for borderless,
                   'auto' to try both

        Returns:
            Path to the created Excel file.
        """
        if output_path is None:
            output_path = str(Path(self.pdf_path).with_suffix('.xlsx'))

        # Extract tables
        if flavor == 'lattice':
            tables = self._extract_tables_lattice(pages)
        elif flavor == 'stream':
            tables = self._extract_tables_stream(pages)
        else:
            tables = self._auto_detect_tables(pages)

        if not tables:
            # Create empty sheet with message
            sheet = self.workbook.create_sheet("No Tables Found")
            sheet.cell(row=1, column=1).value = "No tables were detected in this PDF."
            sheet.cell(row=2, column=1).value = "The PDF may not contain tabular data, or the tables may not be in a recognizable format."
            self.workbook.save(output_path)
            return output_path

        if separate_sheets:
            # Each table in its own sheet
            for idx, table in enumerate(tables):
                sheet_name = f"Table {idx + 1} (Page {table.page})"
                # Excel sheet names have max 31 chars
                sheet_name = sheet_name[:31]
                sheet = self.workbook.create_sheet(sheet_name)
                self._add_table_to_sheet(table, sheet)
        else:
            # All tables in one sheet
            sheet = self.workbook.create_sheet("All Tables")
            current_row = 1

            for idx, table in enumerate(tables):
                # Add table header
                header_cell = sheet.cell(row=current_row, column=1)
                header_cell.value = f"Table {idx + 1} (Page {table.page})"
                header_cell.font = Font(bold=True, size=14)
                current_row += 1

                # Add table
                current_row = self._add_table_to_sheet(table, sheet, current_row)

        # Save workbook
        self.workbook.save(output_path)
        return output_path

    def convert_to_bytes(self, separate_sheets: bool = True,
                         pages: str = 'all', flavor: str = 'auto') -> bytes:
        """
        Convert PDF tables to Excel and return as bytes.

        Returns:
            Excel file as bytes
        """
        # Extract tables
        if flavor == 'lattice':
            tables = self._extract_tables_lattice(pages)
        elif flavor == 'stream':
            tables = self._extract_tables_stream(pages)
        else:
            tables = self._auto_detect_tables(pages)

        if not tables:
            sheet = self.workbook.create_sheet("No Tables Found")
            sheet.cell(row=1, column=1).value = "No tables were detected in this PDF."
        elif separate_sheets:
            for idx, table in enumerate(tables):
                sheet_name = f"Table {idx + 1} (Page {table.page})"[:31]
                sheet = self.workbook.create_sheet(sheet_name)
                self._add_table_to_sheet(table, sheet)
        else:
            sheet = self.workbook.create_sheet("All Tables")
            current_row = 1
            for idx, table in enumerate(tables):
                header_cell = sheet.cell(row=current_row, column=1)
                header_cell.value = f"Table {idx + 1} (Page {table.page})"
                header_cell.font = Font(bold=True, size=14)
                current_row += 1
                current_row = self._add_table_to_sheet(table, sheet, current_row)

        # Save to bytes
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.read()

    def get_table_count(self, pages: str = 'all', flavor: str = 'auto') -> int:
        """Get the number of tables detected in the PDF."""
        if flavor == 'lattice':
            tables = self._extract_tables_lattice(pages)
        elif flavor == 'stream':
            tables = self._extract_tables_stream(pages)
        else:
            tables = self._auto_detect_tables(pages)
        return len(tables)


def convert_pdf_to_excel(input_path: str, output_path: str = None,
                         separate_sheets: bool = True) -> str:
    """
    Convenience function to convert PDF to Excel.

    Args:
        input_path: Path to input PDF file
        output_path: Path to output Excel file (optional)
        separate_sheets: Put each table in a separate sheet

    Returns:
        Path to created Excel file
    """
    converter = PDFToExcelConverter(input_path)
    return converter.convert(output_path, separate_sheets=separate_sheets)


def convert_pdf_bytes_to_excel_bytes(pdf_bytes: bytes,
                                     separate_sheets: bool = True) -> bytes:
    """
    Convert PDF bytes to Excel bytes.

    Args:
        pdf_bytes: PDF file as bytes
        separate_sheets: Put each table in a separate sheet

    Returns:
        Excel file as bytes
    """
    # Write to temp file (Camelot requires file path)
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name

    try:
        converter = PDFToExcelConverter(tmp_path)
        return converter.convert_to_bytes(separate_sheets=separate_sheets)
    finally:
        os.unlink(tmp_path)


# CLI interface
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python pdf_to_excel.py <input.pdf> [output.xlsx]")
        print("Options:")
        print("  --single-sheet    Put all tables in one sheet")
        print("  --lattice         Force lattice mode (bordered tables)")
        print("  --stream          Force stream mode (borderless tables)")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = None
    separate_sheets = True
    flavor = 'auto'

    for arg in sys.argv[2:]:
        if arg == '--single-sheet':
            separate_sheets = False
        elif arg == '--lattice':
            flavor = 'lattice'
        elif arg == '--stream':
            flavor = 'stream'
        elif not arg.startswith('--'):
            output_file = arg

    print(f"Converting {input_file}...")
    converter = PDFToExcelConverter(input_file)
    result = converter.convert(output_file, separate_sheets=separate_sheets, flavor=flavor)
    print(f"Created: {result}")
    print(f"Tables found: {converter.get_table_count()}")
