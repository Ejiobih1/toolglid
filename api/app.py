"""
ConvertX - High-Performance File Conversion API
A production-grade file conversion service similar to ConvertAPI

Architecture:
- FastAPI for high-performance async REST API
- Worker pool for parallel conversions
- Modular converter system (pluggable)
- Temporary file management with auto-cleanup
- Rate limiting and authentication ready

Supported Conversions:
- Documents: PDF ↔ DOCX, DOCX ↔ PDF, HTML → PDF
- Images: PNG ↔ JPG, WebP, resize, compress
- Spreadsheets: XLSX → PDF, CSV → XLSX
- Presentations: PPTX → PDF
"""

import os
import uuid
import asyncio
import tempfile
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from enum import Enum
from contextlib import asynccontextmanager

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Query, Header
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import aiofiles


# ============== Configuration ==============
class Config:
    UPLOAD_DIR = Path(tempfile.gettempdir()) / "convertx_uploads"
    OUTPUT_DIR = Path(tempfile.gettempdir()) / "convertx_outputs"
    MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB
    FILE_RETENTION_HOURS = 24
    MAX_CONCURRENT_CONVERSIONS = 10
    ALLOWED_ORIGINS = ["*"]  # Configure for production
    
    @classmethod
    def ensure_dirs(cls):
        cls.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ============== Models ==============
class ConversionStatus(str, Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"


class ConversionFormat(str, Enum):
    # Documents
    PDF = "pdf"
    DOCX = "docx"
    DOC = "doc"
    HTML = "html"
    TXT = "txt"
    MD = "md"
    RTF = "rtf"
    ODT = "odt"
    
    # Images
    PNG = "png"
    JPG = "jpg"
    JPEG = "jpeg"
    WEBP = "webp"
    GIF = "gif"
    BMP = "bmp"
    TIFF = "tiff"
    SVG = "svg"
    
    # Spreadsheets
    XLSX = "xlsx"
    XLS = "xls"
    CSV = "csv"
    ODS = "ods"
    
    # Presentations
    PPTX = "pptx"
    PPT = "ppt"
    ODP = "odp"


class ConversionRequest(BaseModel):
    source_format: ConversionFormat
    target_format: ConversionFormat
    options: Optional[Dict[str, Any]] = Field(default_factory=dict)


class ConversionResponse(BaseModel):
    job_id: str
    status: ConversionStatus
    source_format: str
    target_format: str
    created_at: datetime
    completed_at: Optional[datetime] = None
    download_url: Optional[str] = None
    error: Optional[str] = None
    file_size: Optional[int] = None
    conversion_time_ms: Optional[int] = None


class ConversionJob(BaseModel):
    job_id: str
    status: ConversionStatus
    source_format: str
    target_format: str
    source_path: Path
    output_path: Optional[Path] = None
    options: Dict[str, Any] = Field(default_factory=dict)
    created_at: datetime = Field(default_factory=datetime.utcnow)
    completed_at: Optional[datetime] = None
    error: Optional[str] = None


# ============== Job Storage (In-Memory for demo, use Redis in production) ==============
class JobStorage:
    def __init__(self):
        self._jobs: Dict[str, ConversionJob] = {}
    
    async def create(self, job: ConversionJob) -> ConversionJob:
        self._jobs[job.job_id] = job
        return job
    
    async def get(self, job_id: str) -> Optional[ConversionJob]:
        return self._jobs.get(job_id)
    
    async def update(self, job_id: str, **kwargs) -> Optional[ConversionJob]:
        if job_id in self._jobs:
            job = self._jobs[job_id]
            for key, value in kwargs.items():
                setattr(job, key, value)
            return job
        return None
    
    async def delete(self, job_id: str) -> bool:
        if job_id in self._jobs:
            del self._jobs[job_id]
            return True
        return False


job_storage = JobStorage()


# ============== Conversion Semaphore ==============
conversion_semaphore = asyncio.Semaphore(Config.MAX_CONCURRENT_CONVERSIONS)


# ============== Converter Registry ==============
class ConverterRegistry:
    """Registry for all supported conversions"""
    
    _converters: Dict[tuple, callable] = {}
    
    @classmethod
    def register(cls, source: ConversionFormat, target: ConversionFormat):
        """Decorator to register a converter function"""
        def decorator(func):
            cls._converters[(source, target)] = func
            return func
        return decorator
    
    @classmethod
    def get_converter(cls, source: ConversionFormat, target: ConversionFormat):
        return cls._converters.get((source, target))
    
    @classmethod
    def get_supported_conversions(cls) -> List[Dict[str, str]]:
        return [
            {"source": src.value, "target": tgt.value}
            for src, tgt in cls._converters.keys()
        ]


# ============== Converters ==============
import subprocess


async def run_command(cmd: List[str], timeout: int = 120) -> tuple:
    """Run a command asynchronously"""
    process = await asyncio.create_subprocess_exec(
        *cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE
    )
    try:
        stdout, stderr = await asyncio.wait_for(
            process.communicate(),
            timeout=timeout
        )
        return process.returncode, stdout.decode(), stderr.decode()
    except asyncio.TimeoutError:
        process.kill()
        raise TimeoutError(f"Command timed out: {' '.join(cmd)}")


# -------- Document Converters (LibreOffice-based) --------
async def libreoffice_convert(input_path: Path, output_format: str, output_dir: Path) -> Path:
    """Convert using LibreOffice in headless mode with optimized settings"""

    cmd = [
        "soffice",
        "--headless",
        "--nofirststartwizard",
        "--norestore",
        "--nologo",
        "--convert-to", output_format,
        "--outdir", str(output_dir),
        str(input_path)
    ]

    # Set environment for better font rendering
    env = os.environ.copy()
    env["SAL_USE_VCLPLUGIN"] = "gen"  # Use generic rendering
    env["HOME"] = "/tmp"  # Ensure LibreOffice has a writable home directory

    process = await asyncio.create_subprocess_exec(
        *cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        env=env
    )

    try:
        stdout, stderr = await asyncio.wait_for(process.communicate(), timeout=180)
        returncode = process.returncode
        stdout_str = stdout.decode()
        stderr_str = stderr.decode()
    except asyncio.TimeoutError:
        process.kill()
        raise RuntimeError("LibreOffice conversion timed out")

    # Log output for debugging
    if stdout_str:
        print(f"LibreOffice stdout: {stdout_str}")
    if stderr_str:
        print(f"LibreOffice stderr: {stderr_str}")

    # Find the output file - check multiple possible names
    expected_output = output_dir / f"{input_path.stem}.{output_format}"
    if expected_output.exists():
        return expected_output

    # Sometimes LibreOffice outputs with different name or case
    for f in output_dir.glob(f"*.{output_format}"):
        return f

    # Also check for lowercase extension
    for f in output_dir.glob(f"*.{output_format.lower()}"):
        return f

    # List all files in output dir for debugging
    all_files = list(output_dir.glob("*"))
    raise RuntimeError(f"Conversion completed but output file not found. Files in output dir: {all_files}")


@ConverterRegistry.register(ConversionFormat.DOCX, ConversionFormat.PDF)
async def docx_to_pdf(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert DOCX to PDF using LibreOffice - auto-detects best export settings"""
    return await libreoffice_convert(input_path, "pdf", output_dir)


@ConverterRegistry.register(ConversionFormat.DOC, ConversionFormat.PDF)
async def doc_to_pdf(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert DOC to PDF using LibreOffice - auto-detects best export settings"""
    return await libreoffice_convert(input_path, "pdf", output_dir)


@ConverterRegistry.register(ConversionFormat.XLSX, ConversionFormat.PDF)
async def xlsx_to_pdf(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert XLSX to PDF using LibreOffice - auto-detects best export settings"""
    # LibreOffice auto-detects the correct filter based on input file type
    return await libreoffice_convert(input_path, "pdf", output_dir)


@ConverterRegistry.register(ConversionFormat.XLS, ConversionFormat.PDF)
async def xls_to_pdf(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert XLS to PDF using LibreOffice - auto-detects best export settings"""
    return await libreoffice_convert(input_path, "pdf", output_dir)


@ConverterRegistry.register(ConversionFormat.PPTX, ConversionFormat.PDF)
async def pptx_to_pdf(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert PPTX to PDF using LibreOffice"""
    return await libreoffice_convert(input_path, "pdf", output_dir)


@ConverterRegistry.register(ConversionFormat.ODT, ConversionFormat.PDF)
async def odt_to_pdf(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert ODT to PDF using LibreOffice"""
    return await libreoffice_convert(input_path, "pdf", output_dir)


@ConverterRegistry.register(ConversionFormat.RTF, ConversionFormat.PDF)
async def rtf_to_pdf(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert RTF to PDF using LibreOffice"""
    return await libreoffice_convert(input_path, "pdf", output_dir)


@ConverterRegistry.register(ConversionFormat.HTML, ConversionFormat.PDF)
async def html_to_pdf(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert HTML to PDF using wkhtmltopdf or LibreOffice"""
    # Try wkhtmltopdf first (better quality)
    output_path = output_dir / f"{input_path.stem}.pdf"
    
    # Options for wkhtmltopdf
    page_size = options.get("page_size", "A4")
    margin_top = options.get("margin_top", "10mm")
    margin_bottom = options.get("margin_bottom", "10mm")
    margin_left = options.get("margin_left", "10mm")
    margin_right = options.get("margin_right", "10mm")
    
    cmd = [
        "wkhtmltopdf",
        "--page-size", page_size,
        "--margin-top", margin_top,
        "--margin-bottom", margin_bottom,
        "--margin-left", margin_left,
        "--margin-right", margin_right,
        "--enable-local-file-access",
        str(input_path),
        str(output_path)
    ]
    
    try:
        returncode, stdout, stderr = await run_command(cmd)
        if returncode == 0 and output_path.exists():
            return output_path
    except Exception:
        pass
    
    # Fallback to LibreOffice
    return await libreoffice_convert(input_path, "pdf", output_dir)


# -------- PDF to Document Converters --------
@ConverterRegistry.register(ConversionFormat.PDF, ConversionFormat.DOCX)
async def pdf_to_docx(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert PDF to DOCX using custom high-quality converter with PyMuPDF"""
    output_path = output_dir / f"{input_path.stem}.docx"

    # Use our custom PDF to Word converter (PyMuPDF + python-docx)
    # This preserves text formatting, images, tables, and layout
    api_dir = str(Path(__file__).parent.absolute())

    script = f'''
import sys
sys.path.insert(0, "{api_dir}")

try:
    from pdf_to_word import PDFToWordConverter

    converter = PDFToWordConverter("{input_path}")
    converter.convert("{output_path}")
    print("SUCCESS")
except ImportError as e:
    # Fallback to pdf2docx if custom converter not available
    print(f"FALLBACK: {{e}}")
    from pdf2docx import Converter
    cv = Converter("{input_path}")
    cv.convert("{output_path}")
    cv.close()
    print("SUCCESS")
except Exception as e:
    import traceback
    print(f"ERROR: {{e}}")
    traceback.print_exc()
    raise
'''

    cmd = ["python3", "-c", script]
    returncode, stdout, stderr = await run_command(cmd, timeout=300)

    if "SUCCESS" in stdout and output_path.exists():
        return output_path

    # Include more detailed error info
    error_msg = stderr if stderr else stdout
    raise RuntimeError(f"PDF to DOCX conversion failed: {error_msg}")


@ConverterRegistry.register(ConversionFormat.PDF, ConversionFormat.TXT)
async def pdf_to_txt(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert PDF to TXT using pdftotext"""
    output_path = output_dir / f"{input_path.stem}.txt"
    
    layout = "-layout" if options.get("preserve_layout", False) else ""
    cmd = ["pdftotext", str(input_path), str(output_path)]
    if layout:
        cmd.insert(1, layout)
    
    returncode, stdout, stderr = await run_command(cmd)
    
    if returncode == 0 and output_path.exists():
        return output_path
    
    raise RuntimeError(f"PDF to TXT conversion failed: {stderr}")


@ConverterRegistry.register(ConversionFormat.PDF, ConversionFormat.PNG)
async def pdf_to_png(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert PDF to PNG using pdftoppm"""
    dpi = options.get("dpi", 150)
    
    cmd = [
        "pdftoppm",
        "-png",
        "-r", str(dpi),
        str(input_path),
        str(output_dir / input_path.stem)
    ]
    
    returncode, stdout, stderr = await run_command(cmd)
    
    if returncode != 0:
        raise RuntimeError(f"PDF to PNG conversion failed: {stderr}")
    
    # Find output files (might be multiple pages)
    output_files = list(output_dir.glob(f"{input_path.stem}*.png"))
    if not output_files:
        raise RuntimeError("Conversion completed but no output files found")
    
    # If single page, return directly; otherwise zip them
    if len(output_files) == 1:
        return output_files[0]
    else:
        # Create a zip of all pages
        import zipfile
        zip_path = output_dir / f"{input_path.stem}_pages.zip"
        with zipfile.ZipFile(zip_path, 'w') as zf:
            for f in sorted(output_files):
                zf.write(f, f.name)
        return zip_path


@ConverterRegistry.register(ConversionFormat.PDF, ConversionFormat.XLSX)
async def pdf_to_xlsx(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert PDF tables to Excel using Camelot"""
    output_path = output_dir / f"{input_path.stem}.xlsx"

    # Use our custom PDF to Excel converter (Camelot + openpyxl)
    api_dir = str(Path(__file__).parent.absolute())

    # Get options
    separate_sheets = options.get("separate_sheets", True)
    flavor = options.get("flavor", "auto")  # lattice, stream, or auto

    script = f'''
import sys
sys.path.insert(0, "{api_dir}")

try:
    from pdf_to_excel import PDFToExcelConverter

    converter = PDFToExcelConverter("{input_path}")
    converter.convert("{output_path}", separate_sheets={separate_sheets}, flavor="{flavor}")
    print("SUCCESS")
except ImportError as e:
    print(f"ERROR: {{e}}")
    raise
except Exception as e:
    import traceback
    print(f"ERROR: {{e}}")
    traceback.print_exc()
    raise
'''

    cmd = ["python3", "-c", script]
    returncode, stdout, stderr = await run_command(cmd, timeout=300)

    if "SUCCESS" in stdout and output_path.exists():
        return output_path

    error_msg = stderr if stderr else stdout
    raise RuntimeError(f"PDF to Excel conversion failed: {error_msg}")


# -------- Image Converters (Pillow/ImageMagick) --------
async def pillow_convert(input_path: Path, output_path: Path, options: Dict) -> Path:
    """Convert images using Pillow"""
    script = f'''
import sys
sys.path.insert(0, '/home/claude/.local/lib/python3.11/site-packages')
from PIL import Image

img = Image.open("{input_path}")

# Handle transparency for formats that don't support it
target_format = "{output_path.suffix[1:].upper()}"
if target_format in ["JPEG", "JPG"] and img.mode in ["RGBA", "P"]:
    # Convert to RGB with white background
    background = Image.new("RGB", img.size, (255, 255, 255))
    if img.mode == "P":
        img = img.convert("RGBA")
    background.paste(img, mask=img.split()[3])
    img = background

# Resize if requested
width = {options.get("width", 0)}
height = {options.get("height", 0)}
if width or height:
    orig_w, orig_h = img.size
    if width and height:
        img = img.resize((width, height), Image.LANCZOS)
    elif width:
        ratio = width / orig_w
        img = img.resize((width, int(orig_h * ratio)), Image.LANCZOS)
    elif height:
        ratio = height / orig_h
        img = img.resize((int(orig_w * ratio), height), Image.LANCZOS)

# Quality for lossy formats
quality = {options.get("quality", 85)}
if target_format in ["JPEG", "JPG", "WEBP"]:
    img.save("{output_path}", quality=quality, optimize=True)
else:
    img.save("{output_path}", optimize=True)

print("SUCCESS")
'''
    
    cmd = ["python3", "-c", script]
    returncode, stdout, stderr = await run_command(cmd)
    
    if "SUCCESS" in stdout and output_path.exists():
        return output_path
    
    raise RuntimeError(f"Image conversion failed: {stderr}")


@ConverterRegistry.register(ConversionFormat.PNG, ConversionFormat.JPG)
async def png_to_jpg(input_path: Path, output_dir: Path, options: Dict) -> Path:
    output_path = output_dir / f"{input_path.stem}.jpg"
    return await pillow_convert(input_path, output_path, options)


@ConverterRegistry.register(ConversionFormat.JPG, ConversionFormat.PNG)
async def jpg_to_png(input_path: Path, output_dir: Path, options: Dict) -> Path:
    output_path = output_dir / f"{input_path.stem}.png"
    return await pillow_convert(input_path, output_path, options)


@ConverterRegistry.register(ConversionFormat.PNG, ConversionFormat.WEBP)
async def png_to_webp(input_path: Path, output_dir: Path, options: Dict) -> Path:
    output_path = output_dir / f"{input_path.stem}.webp"
    return await pillow_convert(input_path, output_path, options)


@ConverterRegistry.register(ConversionFormat.JPG, ConversionFormat.WEBP)
async def jpg_to_webp(input_path: Path, output_dir: Path, options: Dict) -> Path:
    output_path = output_dir / f"{input_path.stem}.webp"
    return await pillow_convert(input_path, output_path, options)


@ConverterRegistry.register(ConversionFormat.WEBP, ConversionFormat.PNG)
async def webp_to_png(input_path: Path, output_dir: Path, options: Dict) -> Path:
    output_path = output_dir / f"{input_path.stem}.png"
    return await pillow_convert(input_path, output_path, options)


@ConverterRegistry.register(ConversionFormat.WEBP, ConversionFormat.JPG)
async def webp_to_jpg(input_path: Path, output_dir: Path, options: Dict) -> Path:
    output_path = output_dir / f"{input_path.stem}.jpg"
    return await pillow_convert(input_path, output_path, options)


@ConverterRegistry.register(ConversionFormat.BMP, ConversionFormat.PNG)
async def bmp_to_png(input_path: Path, output_dir: Path, options: Dict) -> Path:
    output_path = output_dir / f"{input_path.stem}.png"
    return await pillow_convert(input_path, output_path, options)


@ConverterRegistry.register(ConversionFormat.TIFF, ConversionFormat.PNG)
async def tiff_to_png(input_path: Path, output_dir: Path, options: Dict) -> Path:
    output_path = output_dir / f"{input_path.stem}.png"
    return await pillow_convert(input_path, output_path, options)


@ConverterRegistry.register(ConversionFormat.GIF, ConversionFormat.PNG)
async def gif_to_png(input_path: Path, output_dir: Path, options: Dict) -> Path:
    output_path = output_dir / f"{input_path.stem}.png"
    return await pillow_convert(input_path, output_path, options)


# -------- Spreadsheet Converters --------
@ConverterRegistry.register(ConversionFormat.CSV, ConversionFormat.XLSX)
async def csv_to_xlsx(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert CSV to XLSX using openpyxl"""
    output_path = output_dir / f"{input_path.stem}.xlsx"
    
    delimiter = options.get("delimiter", ",")
    
    script = f'''
import sys
sys.path.insert(0, '/home/claude/.local/lib/python3.11/site-packages')
import csv
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

with open("{input_path}", 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter="{delimiter}")
    for row in reader:
        ws.append(row)

wb.save("{output_path}")
print("SUCCESS")
'''
    
    cmd = ["python3", "-c", script]
    returncode, stdout, stderr = await run_command(cmd)
    
    if "SUCCESS" in stdout and output_path.exists():
        return output_path
    
    raise RuntimeError(f"CSV to XLSX conversion failed: {stderr}")


@ConverterRegistry.register(ConversionFormat.XLSX, ConversionFormat.CSV)
async def xlsx_to_csv(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert XLSX to CSV using openpyxl"""
    output_path = output_dir / f"{input_path.stem}.csv"
    
    sheet_index = options.get("sheet", 0)
    
    script = f'''
import sys
sys.path.insert(0, '/home/claude/.local/lib/python3.11/site-packages')
import csv
from openpyxl import load_workbook

wb = load_workbook("{input_path}", data_only=True)
sheets = wb.sheetnames
ws = wb[sheets[{sheet_index}]]

with open("{output_path}", 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)

print("SUCCESS")
'''
    
    cmd = ["python3", "-c", script]
    returncode, stdout, stderr = await run_command(cmd)
    
    if "SUCCESS" in stdout and output_path.exists():
        return output_path
    
    raise RuntimeError(f"XLSX to CSV conversion failed: {stderr}")


# -------- Markdown Converters --------
@ConverterRegistry.register(ConversionFormat.MD, ConversionFormat.PDF)
async def md_to_pdf(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert Markdown to PDF using pandoc"""
    output_path = output_dir / f"{input_path.stem}.pdf"
    
    cmd = [
        "pandoc",
        str(input_path),
        "-o", str(output_path),
        "--pdf-engine=xelatex",
        "-V", "geometry:margin=1in"
    ]
    
    returncode, stdout, stderr = await run_command(cmd)
    
    if returncode == 0 and output_path.exists():
        return output_path
    
    # Fallback: convert to HTML first, then to PDF
    html_path = output_dir / f"{input_path.stem}.html"
    cmd = ["pandoc", str(input_path), "-o", str(html_path), "--standalone"]
    returncode, stdout, stderr = await run_command(cmd)
    
    if returncode == 0 and html_path.exists():
        return await html_to_pdf(html_path, output_dir, options)
    
    raise RuntimeError(f"Markdown to PDF conversion failed: {stderr}")


@ConverterRegistry.register(ConversionFormat.MD, ConversionFormat.HTML)
async def md_to_html(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert Markdown to HTML using pandoc"""
    output_path = output_dir / f"{input_path.stem}.html"
    
    standalone = "--standalone" if options.get("standalone", True) else ""
    cmd = ["pandoc", str(input_path), "-o", str(output_path)]
    if standalone:
        cmd.append(standalone)
    
    returncode, stdout, stderr = await run_command(cmd)
    
    if returncode == 0 and output_path.exists():
        return output_path
    
    raise RuntimeError(f"Markdown to HTML conversion failed: {stderr}")


@ConverterRegistry.register(ConversionFormat.MD, ConversionFormat.DOCX)
async def md_to_docx(input_path: Path, output_dir: Path, options: Dict) -> Path:
    """Convert Markdown to DOCX using pandoc"""
    output_path = output_dir / f"{input_path.stem}.docx"
    
    cmd = ["pandoc", str(input_path), "-o", str(output_path)]
    
    returncode, stdout, stderr = await run_command(cmd)
    
    if returncode == 0 and output_path.exists():
        return output_path
    
    raise RuntimeError(f"Markdown to DOCX conversion failed: {stderr}")


# ============== Conversion Engine ==============
async def process_conversion(job: ConversionJob) -> ConversionJob:
    """Process a conversion job"""
    start_time = datetime.utcnow()
    
    try:
        # Get the appropriate converter
        converter = ConverterRegistry.get_converter(
            ConversionFormat(job.source_format),
            ConversionFormat(job.target_format)
        )
        
        if not converter:
            raise ValueError(f"No converter available for {job.source_format} → {job.target_format}")
        
        # Create output directory for this job
        output_dir = Config.OUTPUT_DIR / job.job_id
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Run conversion with semaphore to limit concurrent conversions
        async with conversion_semaphore:
            output_path = await converter(job.source_path, output_dir, job.options)
        
        # Update job
        job.status = ConversionStatus.COMPLETED
        job.output_path = output_path
        job.completed_at = datetime.utcnow()
        
    except Exception as e:
        job.status = ConversionStatus.FAILED
        job.error = str(e)
        job.completed_at = datetime.utcnow()
    
    return job


# ============== Background Cleanup ==============
async def cleanup_old_files():
    """Remove files older than retention period"""
    cutoff = datetime.utcnow() - timedelta(hours=Config.FILE_RETENTION_HOURS)
    
    for directory in [Config.UPLOAD_DIR, Config.OUTPUT_DIR]:
        if not directory.exists():
            continue
            
        for item in directory.iterdir():
            try:
                mtime = datetime.fromtimestamp(item.stat().st_mtime)
                if mtime < cutoff:
                    if item.is_dir():
                        shutil.rmtree(item)
                    else:
                        item.unlink()
            except Exception:
                pass


# ============== FastAPI Application ==============
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    Config.ensure_dirs()
    yield
    # Shutdown
    pass


app = FastAPI(
    title="ConvertX API",
    description="High-performance file conversion API - Convert documents, images, spreadsheets, and more",
    version="1.0.0",
    lifespan=lifespan
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=Config.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============== API Routes ==============
@app.get("/")
async def root():
    """API information"""
    return {
        "name": "ConvertX API",
        "version": "1.0.0",
        "description": "High-performance file conversion service",
        "documentation": "/docs"
    }


@app.get("/conversions")
async def list_conversions():
    """List all supported conversions"""
    return {
        "supported_conversions": ConverterRegistry.get_supported_conversions(),
        "total": len(ConverterRegistry.get_supported_conversions())
    }


@app.get("/formats")
async def list_formats():
    """List all supported formats"""
    formats = {
        "documents": ["pdf", "docx", "doc", "html", "txt", "md", "rtf", "odt"],
        "images": ["png", "jpg", "jpeg", "webp", "gif", "bmp", "tiff", "svg"],
        "spreadsheets": ["xlsx", "xls", "csv", "ods"],
        "presentations": ["pptx", "ppt", "odp"]
    }
    return formats


@app.post("/convert/{source_format}/to/{target_format}", response_model=ConversionResponse)
async def convert_file(
    source_format: ConversionFormat,
    target_format: ConversionFormat,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    # Optional conversion parameters
    width: Optional[int] = Query(None, description="Image width (for image conversions)"),
    height: Optional[int] = Query(None, description="Image height (for image conversions)"),
    quality: Optional[int] = Query(85, ge=1, le=100, description="Quality for lossy formats (1-100)"),
    dpi: Optional[int] = Query(150, description="DPI for PDF to image conversion"),
    page_size: Optional[str] = Query("A4", description="Page size for HTML to PDF"),
):
    """
    Convert a file from one format to another.
    
    Upload a file and specify the source and target formats.
    The conversion will be processed asynchronously.
    """
    
    # Validate file size
    file.file.seek(0, 2)
    file_size = file.file.tell()
    file.file.seek(0)
    
    if file_size > Config.MAX_FILE_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum size is {Config.MAX_FILE_SIZE / (1024*1024):.0f}MB"
        )
    
    # Check if conversion is supported
    converter = ConverterRegistry.get_converter(source_format, target_format)
    if not converter:
        raise HTTPException(
            status_code=400,
            detail=f"Conversion from {source_format.value} to {target_format.value} is not supported"
        )
    
    # Generate job ID and save file
    job_id = str(uuid.uuid4())
    upload_dir = Config.UPLOAD_DIR / job_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    # Preserve original filename with proper extension
    original_name = file.filename or f"input.{source_format.value}"
    input_path = upload_dir / original_name
    
    async with aiofiles.open(input_path, 'wb') as out_file:
        content = await file.read()
        await out_file.write(content)
    
    # Build options
    options = {}
    if width:
        options["width"] = width
    if height:
        options["height"] = height
    if quality:
        options["quality"] = quality
    if dpi:
        options["dpi"] = dpi
    if page_size:
        options["page_size"] = page_size
    
    # Create job
    job = ConversionJob(
        job_id=job_id,
        status=ConversionStatus.PENDING,
        source_format=source_format.value,
        target_format=target_format.value,
        source_path=input_path,
        options=options
    )
    
    await job_storage.create(job)
    
    # Start conversion in background
    async def run_conversion():
        await job_storage.update(job_id, status=ConversionStatus.PROCESSING)
        result = await process_conversion(job)
        await job_storage.update(
            job_id,
            status=result.status,
            output_path=result.output_path,
            completed_at=result.completed_at,
            error=result.error
        )
    
    background_tasks.add_task(run_conversion)
    
    return ConversionResponse(
        job_id=job_id,
        status=ConversionStatus.PENDING,
        source_format=source_format.value,
        target_format=target_format.value,
        created_at=job.created_at,
        download_url=f"/download/{job_id}"
    )


@app.post("/convert/sync/{source_format}/to/{target_format}")
async def convert_file_sync(
    source_format: ConversionFormat,
    target_format: ConversionFormat,
    file: UploadFile = File(...),
    width: Optional[int] = Query(None),
    height: Optional[int] = Query(None),
    quality: Optional[int] = Query(85, ge=1, le=100),
    dpi: Optional[int] = Query(150),
    page_size: Optional[str] = Query("A4"),
):
    """
    Convert a file synchronously and return the result immediately.
    
    Use this for quick conversions. For large files, use the async endpoint.
    """
    
    # Validate file size (smaller for sync)
    file.file.seek(0, 2)
    file_size = file.file.tell()
    file.file.seek(0)
    
    if file_size > 10 * 1024 * 1024:  # 10MB limit for sync
        raise HTTPException(
            status_code=413,
            detail="File too large for sync conversion. Use the async endpoint or files up to 10MB."
        )
    
    # Check if conversion is supported
    converter = ConverterRegistry.get_converter(source_format, target_format)
    if not converter:
        raise HTTPException(
            status_code=400,
            detail=f"Conversion from {source_format.value} to {target_format.value} is not supported"
        )
    
    # Generate job ID and save file
    job_id = str(uuid.uuid4())
    upload_dir = Config.UPLOAD_DIR / job_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    output_dir = Config.OUTPUT_DIR / job_id
    output_dir.mkdir(parents=True, exist_ok=True)
    
    original_name = file.filename or f"input.{source_format.value}"
    input_path = upload_dir / original_name
    
    async with aiofiles.open(input_path, 'wb') as out_file:
        content = await file.read()
        await out_file.write(content)
    
    # Build options
    options = {}
    if width:
        options["width"] = width
    if height:
        options["height"] = height
    if quality:
        options["quality"] = quality
    if dpi:
        options["dpi"] = dpi
    if page_size:
        options["page_size"] = page_size
    
    try:
        # Run conversion synchronously
        async with conversion_semaphore:
            output_path = await converter(input_path, output_dir, options)
        
        return FileResponse(
            output_path,
            filename=output_path.name,
            media_type="application/octet-stream"
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")
    
    finally:
        # Cleanup will happen via background task
        pass


@app.get("/status/{job_id}", response_model=ConversionResponse)
async def get_job_status(job_id: str):
    """Get the status of a conversion job"""
    job = await job_storage.get(job_id)
    
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    file_size = None
    if job.output_path and job.output_path.exists():
        file_size = job.output_path.stat().st_size
    
    conversion_time = None
    if job.completed_at and job.created_at:
        conversion_time = int((job.completed_at - job.created_at).total_seconds() * 1000)
    
    return ConversionResponse(
        job_id=job.job_id,
        status=job.status,
        source_format=job.source_format,
        target_format=job.target_format,
        created_at=job.created_at,
        completed_at=job.completed_at,
        download_url=f"/download/{job_id}" if job.status == ConversionStatus.COMPLETED else None,
        error=job.error,
        file_size=file_size,
        conversion_time_ms=conversion_time
    )


@app.get("/download/{job_id}")
async def download_result(job_id: str):
    """Download the converted file"""
    job = await job_storage.get(job_id)
    
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.status != ConversionStatus.COMPLETED:
        raise HTTPException(
            status_code=400,
            detail=f"Job is not complete. Current status: {job.status.value}"
        )
    
    if not job.output_path or not job.output_path.exists():
        raise HTTPException(status_code=404, detail="Output file not found")
    
    return FileResponse(
        job.output_path,
        filename=job.output_path.name,
        media_type="application/octet-stream"
    )


@app.post("/cleanup")
async def trigger_cleanup(background_tasks: BackgroundTasks):
    """Trigger cleanup of old files"""
    background_tasks.add_task(cleanup_old_files)
    return {"message": "Cleanup scheduled"}


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat(),
        "max_concurrent_conversions": Config.MAX_CONCURRENT_CONVERSIONS
    }


# ============== Main ==============
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
